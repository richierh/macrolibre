'''
Created on Nov 2, 2018

@author: binakarir
'''
from openpyxl import load_workbook
from sys import stderr

class OpenWorkbook():
    '''
    classdocs
    '''

    def __init__(self, *args):
        '''
        Constructor
        '''
 
        self.Filename = args[0]
        print (self.Filename)
        self.__read__()
        
    def __read__(self):    
        self.wb = load_workbook(filename=self.Filename, data_only =True, keep_links = True)
        self.readsheet = self.wb.get_sheet_names()
        print (self.readsheet)

        '''
        for i in self.readsheet:
            print (i)
        '''
    
    def ChangeTitle(self):
        self.ws = self.wb.get_sheet_by_name(self.readsheet[2])
        print (self.ws.title)
        try:
            self.ws.title="Proses Kehadiran"
            print (self.ws.title)
            self.wb.save(filename="sistempayroll.xlsx")
        except stderr :
            pass
        
index_kerja = 4
hari_kerja="5hk"
bulan = "Agustus"
lembur_jam_ke = "3"
setengah_hari_kerja = "no"
hari_kerja_full = "yes"
hari_libur_kerja = "no"


index_lembur_5hari = {1:1.5,2:3.5,3:5.5,4:7.5,5:9.5,6:11.5,7:13.5,\
                      8:15.5,9:17.5,10:19.5,11:21.5}

 
index =  index_lembur_5hari
print (index.get(index_kerja))


        

if __name__=="__main__":
    Filename = "sistempayroll.xlsx"
    #run = OpenWorkbook(Filename)
    
    